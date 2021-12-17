# -*- coding: utf-8 -*-
"""
-------------------------------------------------
# @Project  :ESS_API
# @File     :测试数据
# @Date     :2021/8/16 15:52
# @Author   :LiuFei
# @Email    :fei.liu@hxgroup.com
# @Software :PyCharm
-------------------------------------------------
"""
"""
一次性获取data目录下所有文件的所有数据
数据存储形式为列表嵌套字典，每一行数据存储为字典。
"""
from openpyxl import load_workbook

from lib.out_log import OutLog

import sys

logger = OutLog().out_log()


class GetRequestData():

    def get_request_data(self):
        logger.info('开始获取data目录下所有Excel文件测试数据')
        test_data = []
        file_name = r'G:\Pycharm_project\flask_demo\data\data1.xlsx'
        wb = load_workbook(file_name)
        all_sheet_name = wb.get_sheet_names()
        for sheet in all_sheet_name:
            ws = wb[sheet]
            for i in range(2, ws.max_row + 1):
                sub_data = {}
                sub_data['case_id'] = ws.cell(i, 1).value
                sub_data['case_name'] = ws.cell(i, 2).value
                sub_data['path_info'] = ws.cell(i, 3).value
                sub_data['method'] = ws.cell(i, 4).value
                sub_data['sql_before'] = ws.cell(i, 5).value
                sub_data['parameters'] = ws.cell(i, 6).value
                sub_data['request_expect_result'] = ws.cell(i, 7).value
                # sub_data['request_actual_result'] = ws.cell(i, 8).value
                sub_data['sql_after'] = ws.cell(i, 9).value
                sub_data['database_compare_sql'] = ws.cell(i, 10).value
                sub_data['database_expect_result'] = ws.cell(i, 11).value
                sub_data['expect_status_code'] = ws.cell(i, 12).value
                sub_data['sheet_name'] = sheet
                sub_data['file_name'] = file_name
                test_data.append(sub_data)
        logger.info(f'文件{file_name}的测试数据提取完成')
        return test_data

    def write_back(self, file_name, row, column, result, sheet_name='sheet1'):
        wb = load_workbook(file_name)
        ws = wb[sheet_name]
        ws.cell(row, column).value = result
        wb.save(file_name)

if __name__ == '__main__':
    data = GetRequestData()
    one = data.get_request_data()
    print(one)
    print(len(one))
    print(len(one[0]))
