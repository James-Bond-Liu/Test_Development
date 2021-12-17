# -*- coding: utf-8 -*-
"""
-------------------------------------------------
# @Project  :ESS_API
# @File     :测试数据
# @Date     :2021/8/16 15:52
# @Author   :LiuFei
# @Email    :fei.liu@hxgroup.com
# @Software :PyCharm
-------------------------------------------------
"""
"""
一次性获取data目录下所有文件的所有数据
数据存储形式为列表嵌套字典，每一行数据存储为字典。
"""
from openpyxl import load_workbook
import pandas as pd
from conf.project_path import *
from lib.out_log import OutLog
from lib.read_config import ReadConfig


logger = OutLog().out_log()
class GetRequestData():


    def get_filename(self):  # 获取目录下所有文件的文件名
        try:
            for root, dirs, files in os.walk(excel_data):
                logger.info("路径{}下的所有文件名为{}".format(excel_data, files))
                return files
        except Exception as e:
            logger.error("获取路径{}下的所有文件名失败，详细信息为{}".format(excel_data, e))
            raise e

    def get_request_data(self):
        logger.info('开始获取data目录下所有Excel文件测试数据')
        test_data = []
        filenames = self.get_filename()
        # mode = eval(ReadConfig().readconfig('MODE', 'mode'))
        for file in filenames:
            path1 = os.path.split(os.path.split(os.path.realpath(__file__))[0])[0]
            path2 = os.path.join(path1, 'data', file)
            logger.info('正在获取Excel文件{}的数据'.format(file))
            wb = load_workbook(path2)
            sheet_names = list(pd.read_excel(path2, sheet_name=None))

            for sheet in sheet_names:
                # if sheet in mode.keys():
                ws = wb[sheet]
                    # if mode[sheet] == 'all':
                logger.info('正在获取Excel文件{}的{}sheet表单数据'.format(file, sheet))
                for i in range(2, ws.max_row + 1):
                    sub_data = {}
                    sub_data['case_id'] = ws.cell(i, 1).value
                    sub_data['case_name'] = ws.cell(i, 2).value
                    sub_data['path_info'] = ws.cell(i, 3).value
                    sub_data['method'] = ws.cell(i, 4).value
                    sub_data['sql_before'] = ws.cell(i, 5).value
                    sub_data['parameters'] = ws.cell(i, 6).value
                    sub_data['request_expect_result'] = ws.cell(i, 7).value
                    # sub_data['request_actual_result'] = ws.cell(i, 8).value
                    sub_data['sql_after'] = ws.cell(i, 9).value
                    sub_data['database_compare_sql'] = ws.cell(i, 10).value
                    sub_data['database_expect_result'] = ws.cell(i, 11).value
                    sub_data['expect_status_code'] = ws.cell(i, 12).value
                    sub_data['sheet_name'] = sheet
                    sub_data['file_name'] = file
                    test_data.append(sub_data)
                # else:
                #     for case_id in mode[sheet]:
                #         sub_data = {}
                #         sub_data['case_id'] = ws.cell(case_id+1, 1).value
                #         sub_data['case_name'] = ws.cell(case_id+1, 2).value
                #         sub_data['path_info'] = ws.cell(case_id+1, 3).value
                #         sub_data['method'] = ws.cell(case_id+1, 4).value
                #         sub_data['sql_before'] = ws.cell(case_id+1, 5).value
                #         sub_data['parameters'] = ws.cell(case_id+1, 6).value
                #         sub_data['request_expect_result'] = ws.cell(case_id+1, 7).value
                #         # sub_data['request_actual_result'] = ws.cell(case_id+1, 8).value
                #         sub_data['sql_after'] = ws.cell(case_id+1, 9).value
                #         sub_data['database_compare_sql'] = ws.cell(case_id+1, 10).value
                #         sub_data['database_expect_result'] = ws.cell(case_id+1, 11).value
                #         sub_data['expect_status_code'] = ws.cell(case_id+1, 12).value
                #         sub_data['sheet_name'] = sheet
                #         sub_data['file_name'] = file
                #         test_data.append(sub_data)

        logger.info('获取data目录下所有Excel文件测试数据完成')
        return test_data


    def write_back(self, file_name, sheet_name, row, column, result):
        wb = load_workbook(file_name)
        ws = wb[sheet_name]
        ws.cell(row, column).value = result
        wb.save(file_name)




if __name__ == '__main__':

    data = GetRequestData()
    one = data.get_request_data()
    print(one)
    print('*'*100)
    print(len(one))

