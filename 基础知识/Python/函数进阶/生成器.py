# -*- coding: utf-8 -*-
# @Time :2021/7/9 12:58
# @Author : liufei
# @File :生成器.PY

"""生成器在Python中，使用了yield 的函数被称为生成器（generator）。

跟普通函数不同的是，生成器是一个返回迭代器的函数，只能用于迭代操作，更简单点理解生成器就是一个迭代器。

在调用生成器运行的过程中，每次遇到yield 时函数会暂停并保存当前所有的运行信息，返回yield 的值, 并在下一次执行next()
方法时从当前位置继续运行。

调用一个生成器函数，返回的是一个迭代器对象。
"""

# 生成器创建方法一，使用（）推导式
s = (i * i for i in range(4))
print(s)
print(next(s))
print(next(s))
print(next(s))
print(next(s))


# 生成器创建方法二，使用yield函数

def gen():
    for i in range(5):
        print("-----------1-----------")
        se = yield i
        print("----------2----------")
        print(se)
        print("------------3-------")


g = gen()
print(g)
print(next(g))
print(g.send(100))  # send方法与生成器进行交互，send方法具有next()方法作用同时可以给生成器传值
g.close()  # 关闭生成器
g.throw(Exception, 'hello world')  # 在生成器内部主动引发一个异常。（异常类型，异常信息）
print(next(g))

"""homework"""

# 当前有一个文件case.xlsx，设计程序将excel中用例读取到一个生成器中
# 方法一：
from openpyxl import load_workbook
def get_data(filename, sheetname):
    wb = load_workbook(filename)
    ws = wb[sheetname]
    max_hang = ws.max_row
    max_lie = ws.max_column
    for row in range(2, max_hang + 1):
        row_data = {}
        for loc in range(1, max_lie + 1):
            row_data[ws.cell(1, loc).value] = ws.cell(row, loc).value
        yield row_data


data = get_data(r'/基础知识/函数进阶/case.xlsx', sheetname='Sheet1')
print(data)
for i in data:
    print(i)
# print(next(data))
# print(next(data))
# print(next(data))
# print(next(data))
# print(next(data))
# print(next(data))

# 方法二：
def get_data2(file_name, sheet_name):
    wb = load_workbook(file_name)
    ws = wb[sheet_name]
    max_hang = ws.max_row
    max_lie = ws.max_column
    row_data = (
    {ws.cell(1, loc).value: ws.cell(row, loc).value for loc in range(1, max_lie + 1)} for row in range(2, max_hang + 1))
    return row_data
