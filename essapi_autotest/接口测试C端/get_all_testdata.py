# -*- coding: utf-8 -*-
"""
-------------------------------------------------
# @Project  :ESS_API
# @File     :测试数据
# @Date     :2021/8/16 15:52
# @Author   :LiuFei
# @Email    :fei.liu@hxgroup.com
# @Software :PyCharm
-------------------------------------------------
"""
"""
一次性获取data目录下所有文件的所有数据
数据存储形式为列表嵌套字典，每一行数据存储为字典。
"""
from openpyxl import load_workbook
import pandas as pd
from conf.project_path import *
from lib.out_log import OutLog

logger = OutLog().out_log()
class GetRequestData():
    # def __init__(self, file_name):
    #     self.file_name = file_name

    def get_filename(self):  # 获取目录下所有文件的文件名
        try:
            for root, dirs, files in os.walk(excel_data):
                logger.info("路径{}下的所有文件名为{}".format(excel_data, files))
                return files
        except Exception as e:
            logger.error("获取路径{}下的所有文件名失败，详细信息为{}".format(excel_data, e))
            raise e

    def get_request_data(self):
        logger.info('开始获取data目录下所有Excel文件测试数据')
        test_data = []
        filenames = self.get_filename()
        for file in filenames:
            path1 = os.path.split(os.path.split(os.path.realpath(__file__))[0])[0]
            path2 = os.path.join(path1, 'data', file)
            logger.info('正在获取Excel文件{}的数据'.format(file))
            wb = load_workbook(path2)
            sheet_names = list(pd.read_excel(path2, sheet_name=None))
            for sheet in sheet_names:
                logger.info('正在获取Excel文件{}的{}sheet表单数据'.format(file,sheet))
                ws = wb[sheet]
                for i in range(2, ws.max_row + 1):
                    sub_data = {}
                    sub_data['case_id'] = ws.cell(i, 1).value
                    sub_data['case_name'] = ws.cell(i, 2).value
                    sub_data['path_info'] = ws.cell(i, 3).value
                    # if sheet == 'login':
                    #     sub_data['url'] = getattr(GlobalData, 'login_url')+ws.cell(i, 3).value
                    # else:
                    #     sub_data['url'] = getattr(GlobalData, 'api_url') + ws.cell(i, 3).value
                    # sub_data['url'] = getattr(GlobalData, 'url')+ws.cell(i, 3).value
                    sub_data['method'] = ws.cell(i, 4).value
                    sub_data['database'] = ws.cell(i, 5).value
                    sub_data['sql_before'] = ws.cell(i, 6).value
                    sub_data['parameters'] = ws.cell(i, 7).value

                    sub_data['expected_result'] = ws.cell(i, 8).value

                    sub_data['actual_result'] = ws.cell(i, 9).value
                    sub_data['sql_after'] = ws.cell(i, 10).value
                    sub_data['sql_result2'] = ws.cell(i, 11).value
                    sub_data['compare_result'] = ws.cell(i, 12).value

                    sub_data['sheet_name'] = sheet
                    sub_data['file_name'] = file
                    test_data.append(sub_data)
        logger.info('获取data目录下所有Excel文件测试数据完成')
        return test_data


    def write_back(self, file_name, sheet_name, row, column, result):
        wb = load_workbook(file_name)
        ws = wb[sheet_name]
        ws.cell(row, column).value = result
        wb.save(file_name)




if __name__ == '__main__':

    data = GetRequestData()
    one = data.get_request_data()
    print(one)
    print(len(one))
