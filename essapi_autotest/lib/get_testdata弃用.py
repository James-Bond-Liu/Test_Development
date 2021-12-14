# -*- coding: utf-8 -*-
"""
-------------------------------------------------
# @Project  :ESS_API
# @File     :测试数据
# @Date     :2021/8/16 15:52
# @Author   :LiuFei
# @Email    :fei.liu@hxgroup.com
# @Software :PyCharm
-------------------------------------------------
"""
"""
一次性获取一个Excel中所有表单的内容
"""
from openpyxl import load_workbook
import pandas as pd
from conf.global_data import GlobalData
from lib.out_log import OutLog

logger = OutLog().out_log()

class GetRequestData():
    def __init__(self, file_name):
        self.file_name = file_name

    def get_request_data(self):
        logger.info('开始获取测试数据')
        test_data = []
        wb = load_workbook(self.file_name)
        sheet_names = list(pd.read_excel(self.file_name, sheet_name=None))
        for sheet in sheet_names:
            # if sheet == 'login':
                ws = wb[sheet]
                for i in range(2, ws.max_row + 1):
                    sub_data = {}
                    sub_data['case_id'] = ws.cell(i, 1).value
                    sub_data['case_name'] = ws.cell(i, 2).value
                    if sheet == 'login':
                        sub_data['url'] = getattr(GlobalData, 'login_url')+ws.cell(i, 3).value
                    else:
                        sub_data['url'] = getattr(GlobalData, 'api_url') + ws.cell(i, 3).value
                    sub_data['method'] = ws.cell(i, 4).value
                    sub_data['parameters'] = ws.cell(i, 5).value
                    sub_data['sql'] = ws.cell(i, 6).value
                    sub_data['expected_result'] = ws.cell(i, 7).value
                    sub_data['actual_result'] = ws.cell(i, 8).value
                    sub_data['compare_result'] = ws.cell(i, 9).value
                    sub_data['sheet_name'] = sheet
                    sub_data['file_name'] = self.file_name.split('\\')[-1]
                    test_data.append(sub_data)
            # else:
            #     ws = wb[sheet]
            #     for i in range(2, ws.max_row + 1):
            #         sub_data = {}
            #         sub_data['case_id'] = ws.cell(i, 1).value
            #         sub_data['case_name'] = ws.cell(i, 2).value
            #         sub_data['url'] = getattr(GlobalData, 'api_url')+ws.cell(i, 3).value
            #         sub_data['method'] = ws.cell(i, 4).value
            #         sub_data['parameters'] = ws.cell(i, 5).value
            #         sub_data['expected_result'] = ws.cell(i, 6).value
            #         sub_data['actual_result'] = ws.cell(i, 7).value
            #         sub_data['compare_result'] = ws.cell(i, 8).value
            #         sub_data['sheet_name'] = sheet
            #         sub_data['file_name'] = self.file_name.split('\\')[-1]
            #         test_data.append(sub_data)
        logger.info('测试数据获取完成')
        return test_data



    def write_back(self, sheet_name, row, column, result):
        wb = load_workbook(self.file_name)
        ws = wb[sheet_name]
        ws.cell(row, column).value = result
        wb.save(self.file_name)




if __name__ == '__main__':
    from conf.project_path import *

    data = GetRequestData(excel_data)
    one = data.get_request_data()
    print(one)
    print(len(one))
